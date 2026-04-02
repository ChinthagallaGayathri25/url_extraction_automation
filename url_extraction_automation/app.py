import streamlit as st
import re
import zipfile
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ===============================
# URL extraction logic
# ===============================

URL_PATTERN = re.compile(r'https?://\S+')

def match_token(line):
    cleaned = line.replace(" ", "").lower()
    return "loadaccesstoken" in cleaned

def extract_urls(lines):
    urls = []

    for i in range(len(lines)):
        line = lines[i]

        if match_token(line):
            urls.extend(URL_PATTERN.findall(line))
            if i + 1 < len(lines):
                urls.extend(URL_PATTERN.findall(lines[i + 1]))

    final = []
    for url in urls:
        url = url.rstrip('\'",)')
        if url not in final:
            final.append(url)

    return final


# ===============================
# Streamlit UI
# ===============================

st.set_page_config(
    page_title="Log File URL Extractor",
    layout="wide",
    page_icon="🔍"
)

st.title("🔍 Log File URL Extractor")
st.write("Upload log/text files and extract URLs containing **loadAccessToken**")

uploaded_files = st.file_uploader(
    "Upload .txt or .log files",
    type=["txt", "log"],
    accept_multiple_files=True
)

if uploaded_files:
    st.success(f"{len(uploaded_files)} file(s) uploaded")

    if st.button("Extract URLs"):

        results = {}
        combined_rows = []
        summary_rows = []

        # ===============================
        # Process files
        # ===============================
        for file in uploaded_files:
            content = file.read().decode("utf-8", errors="ignore")
            lines = content.splitlines()
            urls = extract_urls(lines)

            results[file.name] = urls

            summary_rows.append({
                "File Name": file.name,
                "URL Count": len(urls)
            })

            numbered_urls = "\n".join(
                [f"{i+1}. {u}" for i, u in enumerate(urls)]
            )

            combined_rows.append({
                "File Name": file.name,
                "URLs": numbered_urls
            })

        # ===============================
        # ✅ UI DISPLAY (RESTORED)
        # ===============================
        st.header("📄 Extracted URLs by File")

        for filename, urls in results.items():
            with st.expander(f"📁 {filename} — {len(urls)} URL(s)", expanded=True):
                if urls:
                    for u in urls:
                        st.code(u)
                else:
                    st.info("No URLs found")

        # ===============================
        # Create ZIP with Excel outputs
        # ===============================
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:

            # Per-file Excel
            for filename, urls in results.items():
                df = pd.DataFrame({"URL": urls})
                buf = io.BytesIO()
                df.to_excel(buf, index=False)
                zipf.writestr(f"{filename}.xlsx", buf.getvalue())

            # Combined Excel (formatted)
            combined_df = pd.DataFrame(combined_rows)
            buf = io.BytesIO()
            combined_df.to_excel(buf, index=False)

            buf.seek(0)
            wb = load_workbook(buf)
            ws = wb.active

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 120

            for row in ws.iter_rows(min_row=2):
                cell = row[1]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row[0].row].height = None

            final_buf = io.BytesIO()
            wb.save(final_buf)
            zipf.writestr("combined_urls.xlsx", final_buf.getvalue())

            # Summary Excel
            summary_df = pd.DataFrame(summary_rows)
            buf2 = io.BytesIO()
            summary_df.to_excel(buf2, index=False)
            zipf.writestr("summary.xlsx", buf2.getvalue())

        st.header("⬇️ Download Results")
        st.download_button(
            "Download ZIP (Excel files)",
            zip_buffer.getvalue(),
            file_name="url_extraction_results.zip",
            mime="application/zip"
        )

st.sidebar.title("ℹ️ How This App Works")

st.sidebar.markdown("""
### Step 1: Upload Log Files
- Upload one or more `.txt` or `.log` files using the upload button.
- Each file is processed independently.

### Step 2: Extract URLs
- Click the **Extract URLs** button.
- The app scans the logs and finds URLs containing `loadAccessToken`.

### Step 3: View Extracted URLs
- Extracted URLs are displayed file‑wise on the screen.
- You can expand each file section to view all URLs.

### Step 4: Download Output
- Click **Download ZIP** to get the results.
- The ZIP file contains:
    - One Excel file per uploaded log
    - A combined Excel file (URLs grouped by file)
    - A summary Excel file (URL count per file)
""")
